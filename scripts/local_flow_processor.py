#!/usr/bin/env python3
"""ISAAI — Local Flow Processor (Xetra Reports).

Simulates what the Power Automate flow does:
  1. Reads the XML report files from each run
  2. Saves them as styled XLSX files (the main deliverable)
  3. Logs the processing result to a local CSV (SharePoint equivalent)
  4. Sends a verification summary to stdout

No validation logic is performed — the status values are read directly
from the XML report's <Summary>/<OverallStatus> element.
"""

import csv
import sys
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Error: openpyxl is required. Install: pip install openpyxl")
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[1]
RUNS_DIR = ROOT / "tests" / "runs"
LOG_FILE = ROOT / "tests" / "Report_Processing_Log_Local.csv"

# Styles
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="00335B", end_color="00335B", fill_type="solid")
PASS_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
VIOL_FILL = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
PASS_FONT = Font(name="Calibri", color="2E7D32")
VIOL_FONT = Font(name="Calibri", color="D9383A", bold=True)
BODY_FONT = Font(name="Calibri", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def xml_to_xlsx_daily(xml_path, xlsx_path):
    """Convert Report Type A (daily trades) from XML to styled XLSX."""
    tree = ET.parse(xml_path)
    root = tree.getroot()

    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Xetra Trades"

    # Metadata sheet header
    meta = root.find("Metadata")
    report_id = meta.findtext("ReportID", "")
    report_date = meta.findtext("CoveragePeriod", "")
    venue = meta.findtext("TradingVenue", "XETR")
    member = meta.findtext("MemberID", "")

    # Trade data headers
    headers = [
        "TradeID", "ExecutionTime", "ISIN", "Instrument", "Side",
        "Quantity", "Price", "Currency", "OrderType", "TraderID",
        "DayViol", "ViolationDetails", "Status"
    ]

    # Write metadata rows
    ws.append([f"Xetra Daily Trading Report — {report_date}"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="00335B")
    ws.append([f"Report ID: {report_id}  |  Venue: {venue}  |  Member: {member}"])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws["A2"].font = Font(name="Calibri", size=10, color="666666")
    ws.append([])  # blank row

    # Headers
    ws.append(headers)
    header_row = 4
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    # Trade rows
    trades = root.findall(".//Trade")
    for trade in trades:
        row_data = [trade.findtext(h, "") for h in headers]
        ws.append(row_data)
        row_num = ws.max_row
        status = trade.findtext("Status", "Pass")
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            if status == "Violation":
                cell.fill = VIOL_FILL
                if col_idx in (11, 12, 13):  # DayViol, ViolationDetails, Status
                    cell.font = VIOL_FONT
            else:
                cell.fill = PASS_FILL

    # Summary section
    summary = root.find("Summary")
    ws.append([])
    ws.append(["Summary"])
    ws.cell(row=ws.max_row, column=1).font = Font(name="Calibri", bold=True, size=12, color="00335B")
    ws.append([f"Total Trades: {summary.findtext('TotalTrades', '0')}"])
    ws.append([f"Total Volume: EUR {summary.findtext('TotalVolume', '0')}"])
    ws.append([f"Violations Found: {summary.findtext('ViolationsFound', '0')}"])
    ws.append([f"Overall Status: {summary.findtext('OverallStatus', 'Pass')}"])

    # Auto-width columns
    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[ws.cell(row=header_row, column=col_idx).column_letter].width = max(14, len(header) + 4)
    ws.column_dimensions["L"].width = 50  # ViolationDetails

    wb.save(xlsx_path)


def xml_to_xlsx_monthly(xml_path, xlsx_path):
    """Convert Report Type B (monthly trader summary) from XML to styled XLSX."""
    tree = ET.parse(xml_path)
    root = tree.getroot()

    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Trader Summary"

    meta = root.find("Metadata")
    report_id = meta.findtext("ReportID", "")
    coverage = meta.findtext("CoveragePeriod", "")
    venue = meta.findtext("TradingVenue", "XETR")
    member = meta.findtext("MemberID", "")

    headers = [
        "TraderID", "TraderName", "Desk", "TotalTrades", "TotalVolume",
        "AvgTradeSize", "TopInstrument", "MonthViol", "ViolationCount",
        "ViolationDetails", "ComplianceRating", "Status"
    ]

    # Metadata
    ws.append([f"Xetra Monthly Trader Summary — {coverage}"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="00335B")
    ws.append([f"Report ID: {report_id}  |  Venue: {venue}  |  Member: {member}"])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws["A2"].font = Font(name="Calibri", size=10, color="666666")
    ws.append([])

    # Headers
    ws.append(headers)
    header_row = 4
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    # Trader rows
    traders = root.findall(".//Trader")
    for trader in traders:
        row_data = [trader.findtext(h, "") for h in headers]
        ws.append(row_data)
        row_num = ws.max_row
        status = trader.findtext("Status", "Pass")
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            if status == "Violation":
                cell.fill = VIOL_FILL
                if col_idx in (8, 9, 10, 11, 12):
                    cell.font = VIOL_FONT
            else:
                cell.fill = PASS_FILL

    # Summary
    summary = root.find("Summary")
    ws.append([])
    ws.append(["Summary"])
    ws.cell(row=ws.max_row, column=1).font = Font(name="Calibri", bold=True, size=12, color="00335B")
    ws.append([f"Total Traders: {summary.findtext('TotalTraders', '0')}"])
    ws.append([f"Total Trades: {summary.findtext('TotalTrades', '0')}"])
    ws.append([f"Total Volume: EUR {summary.findtext('TotalVolume', '0')}"])
    ws.append([f"Violations Found: {summary.findtext('ViolationsFound', '0')}"])
    ws.append([f"Overall Status: {summary.findtext('OverallStatus', 'Pass')}"])

    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[ws.cell(row=header_row, column=col_idx).column_letter].width = max(14, len(header) + 4)
    ws.column_dimensions["G"].width = 35  # TopInstrument
    ws.column_dimensions["J"].width = 50  # ViolationDetails

    wb.save(xlsx_path)


def process_all_runs():
    """Process all generated runs."""
    run_dirs = sorted(RUNS_DIR.glob("run_*"), key=lambda p: int(p.name.split("_")[1]))
    if not run_dirs:
        print("No runs found. Run scripts/simulate_runs.py first.")
        sys.exit(1)

    print(f"Processing {len(run_dirs)} Xetra report runs...\n")

    log_rows = []

    for run_dir in run_dirs:
        # Find XML files
        xml_a = list(run_dir.glob("*_ReportTypeA.xml"))
        xml_b = list(run_dir.glob("*_ReportTypeB.xml"))

        if not xml_a or not xml_b:
            print(f"  Skipping {run_dir.name}: missing XML files")
            continue

        xml_a = xml_a[0]
        xml_b = xml_b[0]
        date_str = xml_a.stem.split("_")[0]

        # Convert XML → XLSX
        xlsx_a = run_dir / "evidence_ReportTypeA.xlsx"
        xlsx_b = run_dir / "evidence_ReportTypeB.xlsx"
        xml_to_xlsx_daily(xml_a, xlsx_a)
        xml_to_xlsx_monthly(xml_b, xlsx_b)

        # Read statuses from XML
        tree_a = ET.parse(xml_a)
        tree_b = ET.parse(xml_b)
        status_a = tree_a.find(".//Summary/OverallStatus").text
        status_b = tree_b.find(".//Summary/OverallStatus").text
        viols_a = tree_a.find(".//Summary/ViolationsFound").text
        viols_b = tree_b.find(".//Summary/ViolationsFound").text

        has_violation = (status_a == "Violation" or status_b == "Violation")
        overall = "Exception" if has_violation else "Completed"
        flag = "Yes" if has_violation else "No"

        detail_a = f"{viols_a} trade violation(s) flagged" if status_a == "Violation" else "No violations"
        detail_b = f"{viols_b} trader violation(s) flagged" if status_b == "Violation" else "No violations"

        marker = "⚠️" if has_violation else "✅"
        print(f"  {marker} {run_dir.name} [{date_str}]: A={status_a} | B={status_b} → {overall}")

        log_rows.append({
            "Title": f"Xetra Report Ingestion — {date_str}",
            "ReceivedDateTime": datetime.now().isoformat() + "Z",
            "SenderEmail": "xetra-reporting@fra-uas.de",
            "ReportDate": date_str,
            "ReportA_Status": status_a,
            "ReportB_Status": status_b,
            "ReportA_ViolationDetails": detail_a,
            "ReportB_ViolationDetails": detail_b,
            "OverallStatus": overall,
            "ExceptionFlag": flag,
            "EvidenceA_Link": str(xlsx_a.relative_to(ROOT)),
            "EvidenceB_Link": str(xlsx_b.relative_to(ROOT)),
        })

    # Write local SharePoint log
    fieldnames = list(log_rows[0].keys())
    with open(LOG_FILE, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(log_rows)

    completed = sum(1 for r in log_rows if r["OverallStatus"] == "Completed")
    exceptions = sum(1 for r in log_rows if r["OverallStatus"] == "Exception")

    print(f"\nProcessing complete: {len(log_rows)} runs")
    print(f"  ✅ Completed: {completed}")
    print(f"  ⚠️  Exception: {exceptions}")
    print(f"Local log: {LOG_FILE.relative_to(ROOT)}")
    print(f"XLSX evidence files generated in each run directory.")


if __name__ == "__main__":
    process_all_runs()
